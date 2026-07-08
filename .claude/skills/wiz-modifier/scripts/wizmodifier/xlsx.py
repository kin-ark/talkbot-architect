"""Excel (.xlsx) I/O for WIZ KB/intent sheets.

WIZ exports are named .xls but are xlsx bytes; openpyxl rejects the .xls
extension, so we load from BytesIO (content, not name).
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl


def read_rows(path: str | Path, sheet: str | None = None) -> list[list]:
    """Return all rows (as lists, values-only) of the named sheet (or the first)."""
    p = Path(path)
    try:
        data = p.read_bytes()
    except OSError as e:
        raise ValueError(f"cannot read xlsx {p}: {e}") from e
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # openpyxl raises several types on bad input
        raise ValueError(f"not a readable xlsx: {p} ({e})") from e
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise ValueError(f"sheet {sheet!r} not in {wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def write_sheet(
    path: str | Path, header: list[str], rows: list[list], note: str | None = None
) -> None:
    """Write a single-sheet xlsx: optional note in A1, then header, then rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    if note is not None:
        ws.append([note])
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    wb.save(str(path))
